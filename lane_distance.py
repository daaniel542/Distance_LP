# lane_distance.py
#!/usr/bin/env python3
"""
Lane Distance Calculator (Mapbox Edition)
-----------------------------------------
Calculates crow-flight distances between origin and destination cities,
automatically picking the farthest match when names collide, and flagging
ambiguous lookups.
"""

import os
import sys
import re
import argparse
import logging
import pathlib
import sqlite3
from typing import Optional, Tuple, List, Callable, Dict, Union

import pandas as pd
from mapbox import Geocoder
from geopy.distance import geodesic
from geopy.extra.rate_limiter import RateLimiter
from geopy.exc import GeocoderUnavailable

DEFAULT_CACHE_DB = pathlib.Path("city_cache.sqlite")
EXPECTED_COLS = ["origin", "destination", "lane_distance_mi"]

# --------------------------------------------------
def open_cache(db_path: Union[str, pathlib.Path] = DEFAULT_CACHE_DB
) -> sqlite3.Connection:
    """Open or create the SQLite cache."""
    try:
        conn = sqlite3.connect(db_path)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS coords (
                place TEXT PRIMARY KEY,
                lat   REAL,
                lon   REAL
            )
        """)
        return conn
    except sqlite3.OperationalError as e:
        sys.exit(f"❌ Cannot open cache DB '{db_path}': {e}")

# --------------------------------------------------
def clean_place(raw: object) -> Optional[str]:
    """Standardize place names: trim, uppercase, remove punctuation."""
    if pd.isna(raw):
        return None
    s = re.sub(r"\s+", " ", str(raw).strip()).upper().strip(".,;:")
    return s or None

# --------------------------------------------------
def make_mapbox_geocoder() -> Callable[[str], List[object]]:
    """Build a RateLimiter-wrapped Mapbox geocoder returning all candidates."""
    token = os.getenv("MAPBOX_TOKEN")
    if not token:
        sys.exit("❌ Please set your Mapbox token via export MAPBOX_TOKEN=...")
    mb = Geocoder(access_token=token)

    def _forward_all(place: str):
        try:
            resp = mb.forward(place, limit=5)
            data = resp.geojson()
            features = data.get("features") or []
            class Loc: pass
            locs: List[Loc] = []
            for feat in features:
                lon, lat = feat["geometry"]["coordinates"]
                loc = Loc()
                loc.latitude = lat
                loc.longitude = lon
                locs.append(loc)
            return locs
        except GeocoderUnavailable as e:
            logging.warning(f"[geocode] service unavailable: {e}")
            return []

    return RateLimiter(_forward_all, min_delay_seconds=1, max_retries=2, error_wait_seconds=5)

# --------------------------------------------------
def geocode_place(
    place: Optional[str],
    geocode_func: Callable[[str], Optional[object]],
    cache_conn: sqlite3.Connection
) -> Optional[Tuple[float, float]]:
    """Geocode place from cache or via Mapbox."""
    if place is None:
        return None

    cache_conn.execute("""
        CREATE TABLE IF NOT EXISTS coords (
            place TEXT PRIMARY KEY,
            lat   REAL,
            lon   REAL
        )
    """)

    cur = cache_conn.execute("SELECT lat, lon FROM coords WHERE place = ?", (place,))
    row = cur.fetchone()
    if row:
        logging.debug(f"[cache] {place} → {row}")
        return row

    try:
        loc = geocode_func(place)
    except GeocoderUnavailable as e:
        logging.warning(f"[geocode] service unavailable: {e}")
        return None

    if loc is None:
        logging.warning(f"[geocode] no result for '{place}'")
        return None

    cache_conn.execute(
        "INSERT OR REPLACE INTO coords(place, lat, lon) VALUES (?,?,?)",
        (place, loc.latitude, loc.longitude)
    )
    cache_conn.commit()
    logging.debug(f"[cache] saved {place} → ({loc.latitude}, {loc.longitude})")
    return (loc.latitude, loc.longitude)

# --------------------------------------------------
def distance_miles(
    p1: Optional[Tuple[float, float]],
    p2: Optional[Tuple[float, float]]
) -> float:
    """Calculate geodesic distance or return NaN."""
    if p1 is None or p2 is None:
        return float("nan")
    return geodesic(p1, p2).miles

# --------------------------------------------------
def calculate_distances_and_flags(
    df: pd.DataFrame,
    geocode_func: Callable[[str], List[object]],
    cache_conn: sqlite3.Connection
) -> Tuple[List[float], List[bool], List[bool]]:
    """
    For each row, geocode all candidates for origin and destination,
    pick the pair that gives the max distance, and flag ambiguous cells.
    """
    dists: List[float] = []
    origin_amb: List[bool] = []
    dest_amb:   List[bool] = []

    for o_raw, d_raw in zip(df["origin"], df["destination"]):
        o_locs = geocode_func(o_raw) or []
        d_locs = geocode_func(d_raw) or []

        o_multi = len(o_locs) > 1
        d_multi = len(d_locs) > 1

        if not o_locs or not d_locs:
            dists.append(float("nan"))
            origin_amb.append(o_multi)
            dest_amb.append(d_multi)
            continue

        best = -1.0
        for o_loc in o_locs:
            for d_loc in d_locs:
                dist = geodesic(
                    (o_loc.latitude, o_loc.longitude),
                    (d_loc.latitude, d_loc.longitude)
                ).miles
                if dist > best:
                    best = dist

        dists.append(round(best, 1))
        origin_amb.append(o_multi)
        dest_amb.append(d_multi)

    return dists, origin_amb, dest_amb

# --------------------------------------------------
def calculate_distances(
    origins: List[str],
    destinations: List[str],
    geocode_func: Callable[[str], Optional[object]],
    cache_conn: sqlite3.Connection
) -> List[float]:
    """Core logic: map origin/destination → geocode → compute distance."""
    coords: Dict[str, Optional[Tuple[float, float]]] = {}
    for place in set(origins) | set(destinations):
        coords[place] = geocode_place(place, geocode_func, cache_conn)

    return [
        distance_miles(coords.get(o), coords.get(d))
        for o, d in zip(origins, destinations)
    ]

# --------------------------------------------------
def process_file(
    path_in: pathlib.Path,
    path_out: pathlib.Path,
) -> pd.DataFrame:
    """End-to-end file processor: clean, geocode, compute, write."""
    logging.info(f"▶ Reading {path_in}")
    ext = path_in.suffix.lower()
    try:
        df = (
            pd.read_excel(path_in)
            if ext in (".xls", ".xlsx")
            else pd.read_csv(path_in)
        )
    except Exception as e:
        sys.exit(f"❌ Failed to read '{path_in}': {e}")

    if df.shape[1] < 2:
        sys.exit("❌ Need at least 2 columns (origin, destination)")

    df = df.iloc[:, :2]
    df.columns = EXPECTED_COLS[:2]
    df["origin"]      = df["origin"].map(clean_place)
    df["destination"] = df["destination"].map(clean_place)

    before = len(df)
    df.dropna(subset=["origin", "destination"], inplace=True)
    dropped = before - len(df)
    if dropped:
        logging.warning(f"Dropped {dropped} row(s) with blank city names")

    geocode_func = make_mapbox_geocoder()
    cache_conn    = open_cache()

    logging.info("▶ Geocoding & computing distances...")
    dists, origin_amb, dest_amb = calculate_distances_and_flags(
        df, geocode_func, cache_conn
    )
    df["lane_distance_mi"]     = dists
    df["origin_ambiguous"]     = origin_amb
    df["destination_ambiguous"]= dest_amb

    path_out.parent.mkdir(parents=True, exist_ok=True)
    logging.info(f"▶ Writing {path_out}")
    try:
        if ext == ".csv":
            df.to_csv(path_out, index=False)
        else:
            from openpyxl.styles import PatternFill
            with pd.ExcelWriter(path_out, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name="Distances")
                wb = writer.book
                ws = writer.sheets["Distances"]
                yellow = PatternFill(fill_type="solid", fgColor="FFFF00")
                # highlight ambiguous origin (col A) & destination (col B)
                for row_idx, amb in enumerate(origin_amb, start=2):
                    if amb:
                        ws.cell(row=row_idx, column=1).fill = yellow
                for row_idx, amb in enumerate(dest_amb, start=2):
                    if amb:
                        ws.cell(row=row_idx, column=2).fill = yellow
                writer.save()
    except Exception as e:
        sys.exit(f"❌ Failed to write '{path_out}': {e}")

    cache_conn.close()
    return df

# --------------------------------------------------
def main():
    p = argparse.ArgumentParser(
        description="Crow-flight Lane Distance Calculator (Mapbox)"
    )
    p.add_argument(
        "infile", type=pathlib.Path,
        help="Input file (CSV or Excel)"
    )
    p.add_argument(
        "-o", "--out", type=pathlib.Path,
        default=pathlib.Path("lane_output.csv"),
        help="Output file (default: lane_output.csv)"
    )
    p.add_argument(
        "-v", "--verbose", action="store_true",
        help="Enable verbose logging"
    )
    args = p.parse_args()

    log_level = logging.DEBUG if args.verbose else logging.INFO
    logging.basicConfig(
        level=log_level,
        format="%(asctime)s %(levelname)s | %(message)s"
    )

    if not args.infile.exists():
        sys.exit(f"❌ Input file not found: {args.infile}")

    try:
        process_file(args.infile, args.out)
        logging.info("✅ Done.")
    except KeyboardInterrupt:
        logging.error("❌ Interrupted.")
        sys.exit(1)

if __name__ == "__main__":
    main()
